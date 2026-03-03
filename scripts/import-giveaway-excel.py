#!/usr/bin/env python3
"""
Import Golf VX Anniversary Giveaway applications from Excel into the database.
Uses the second sheet which has 45 real entries with richer data.
Upserts by email to avoid duplicates.
"""
import openpyxl
import mysql.connector
import os
import sys
from datetime import datetime, date, time

EXCEL_PATH = "/home/ubuntu/upload/GOLFVXAHAnniversaryApplications2026.xlsx"
DATABASE_URL = os.environ.get("DATABASE_URL", "")

def parse_db_url(url):
    """Parse mysql://user:pass@host:port/db into connection params."""
    # Format: mysql://user:pass@host:port/dbname?params
    url = url.replace("mysql://", "").replace("mysql2://", "")
    if "?" in url:
        url = url.split("?")[0]
    user_pass, rest = url.split("@", 1)
    if ":" in user_pass:
        user, password = user_pass.split(":", 1)
    else:
        user, password = user_pass, ""
    if "/" in rest:
        host_port, dbname = rest.rsplit("/", 1)
    else:
        host_port, dbname = rest, ""
    if ":" in host_port:
        host, port = host_port.rsplit(":", 1)
        port = int(port)
    else:
        host, port = host_port, 3306
    return {"host": host, "port": port, "user": user, "password": password, "database": dbname}

def normalize_phone(phone):
    """Normalize phone number to string."""
    if phone is None:
        return None
    return str(int(phone)) if isinstance(phone, float) else str(phone)

def normalize_bool(val):
    """Normalize boolean values."""
    if val is None:
        return False
    if isinstance(val, bool):
        return val
    if isinstance(val, str):
        return val.strip().lower() in ("yes", "true", "1", "y")
    return bool(val)

def normalize_timestamp(val):
    """Convert datetime to MySQL-compatible string."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(val, date):
        return datetime(val.year, val.month, val.day).strftime("%Y-%m-%d %H:%M:%S")
    return str(val)

def main():
    print(f"Loading Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH)
    
    # Use the second sheet which has 45 real entries
    ws = wb["Annual Anniversary Application"]
    print(f"Sheet: {ws.title}, rows: {ws.max_row}, cols: {ws.max_column}")
    
    # Row 2 has actual headers
    headers = [ws.cell(2, c).value for c in range(1, ws.max_column+1)]
    print(f"Headers: {headers[:15]}...")
    
    # Column indices (0-based from headers list)
    # Col 1=timestamp, 2=Name, 3=Email, 4=Phone, 5=AgeRange, 6=Gender, 7=City
    # 8=IllinoisResident, 9=GolfExperience, 10=VisitedBefore, 11=FirstHeard
    # 12=FirstVisitTiming, 13=IndoorGolfFreq, 14=WhatStoodOut, 15=IndoorGolfFamiliarity
    # 16=GolfVXInterest, 25=LoveOfGame, 26=HelpGrowCommunity, 27=ConnectedShare
    # 30=SocialHandle, 31=GroupsEngaged, 32=BestTimeToCall, 33=HeardAboutOffer
    
    # Connect to DB
    db_params = parse_db_url(DATABASE_URL)
    print(f"Connecting to DB: {db_params['host']}:{db_params['port']}/{db_params['database']}")
    conn = mysql.connector.connect(**db_params, ssl_disabled=False)
    cursor = conn.cursor(dictionary=True)
    
    # Get existing emails to check for duplicates
    cursor.execute("SELECT email, id, submissionTimestamp FROM giveawayApplications WHERE isTestEntry = 0")
    existing = {row["email"]: row for row in cursor.fetchall()}
    print(f"Existing real entries in DB: {len(existing)}")
    
    inserted = 0
    updated = 0
    skipped = 0
    
    for row_idx in range(3, ws.max_row + 1):
        row = [ws.cell(row_idx, c).value for c in range(1, ws.max_column + 1)]
        
        # Skip empty rows
        if not row[1]:  # Name is empty
            continue
        
        # Skip test entries
        name = str(row[1]).strip()
        if "TEST" in name.upper():
            skipped += 1
            continue
        
        timestamp = normalize_timestamp(row[0])
        email = str(row[2]).strip().lower() if row[2] else None
        if not email:
            skipped += 1
            continue
        
        phone = normalize_phone(row[3])
        age_range = str(row[4]).strip() if row[4] else None
        gender = str(row[5]).strip() if row[5] else None
        city = str(row[6]).strip() if row[6] else None
        illinois_resident = normalize_bool(row[7])
        golf_experience = str(row[8]).strip() if row[8] else None
        visited_before = "Existing" if normalize_bool(row[9]) else "New"
        indoor_familiarity = str(row[14]).strip() if row[14] else None
        best_time_to_call = str(row[31]).strip() if row[31] else None
        how_heard = str(row[32]).strip() if row[32] else None
        
        # Build rich notes from passion story + community growth
        notes_parts = []
        if row[24]:  # Love of the Game story
            notes_parts.append(f"Passion Story: {str(row[24]).strip()}")
        if row[25]:  # Help Grow Community
            notes_parts.append(f"Community Growth: {str(row[25]).strip()}")
        if row[29]:  # Social Handle
            notes_parts.append(f"Social: @{str(row[29]).strip()}")
        notes = "\n\n".join(notes_parts) if notes_parts else None
        
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        if email in existing:
            # Update existing entry with richer data
            cursor.execute("""
                UPDATE giveawayApplications SET
                  phone = %s,
                  ageRange = %s,
                  gender = %s,
                  city = %s,
                  illinoisResident = %s,
                  golfExperienceLevel = %s,
                  visitedBefore = %s,
                  indoorGolfFamiliarity = %s,
                  bestTimeToCall = %s,
                  howDidTheyHear = %s,
                  notes = COALESCE(NULLIF(%s, ''), notes),
                  lastSyncedAt = %s,
                  updatedAt = %s
                WHERE email = %s AND isTestEntry = 0
            """, (phone, age_range, gender, city, illinois_resident,
                  golf_experience, visited_before, indoor_familiarity,
                  best_time_to_call, how_heard, notes, now, now, email))
            updated += 1
        else:
            # Insert new entry
            cursor.execute("""
                INSERT INTO giveawayApplications
                  (submissionTimestamp, name, email, phone, ageRange, gender, city,
                   illinoisResident, golfExperienceLevel, visitedBefore, indoorGolfFamiliarity,
                   bestTimeToCall, howDidTheyHear, notes, status, isTestEntry, lastSyncedAt, createdAt, updatedAt)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'pending', 0, %s, %s, %s)
            """, (timestamp, name, email, phone, age_range, gender, city,
                  illinois_resident, golf_experience, visited_before, indoor_familiarity,
                  best_time_to_call, how_heard, notes, now, now, now))
            inserted += 1
    
    conn.commit()
    cursor.close()
    conn.close()
    
    print(f"\n✅ Import complete:")
    print(f"   Inserted: {inserted} new entries")
    print(f"   Updated:  {updated} existing entries")
    print(f"   Skipped:  {skipped} (test/empty)")
    
    # Verify final count
    conn2 = mysql.connector.connect(**db_params)
    c2 = conn2.cursor()
    c2.execute("SELECT COUNT(*) FROM giveawayApplications WHERE isTestEntry = 0")
    total = c2.fetchone()[0]
    c2.execute("SELECT MAX(submissionTimestamp) FROM giveawayApplications WHERE isTestEntry = 0")
    latest = c2.fetchone()[0]
    print(f"   DB total real entries: {total}")
    print(f"   Latest submission: {latest}")
    conn2.close()

if __name__ == "__main__":
    main()
